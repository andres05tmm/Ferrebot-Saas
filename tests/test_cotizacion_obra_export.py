"""Motor de export Excel del cotizador AIU (formato PROVISIONAL): pura, sin BD ni red.

Verifica que `render_cotizacion_excel` produce un `.xlsx` válido (round-trip con openpyxl) y que el
desglose AIU llega a la hoja tal cual lo calculó la función pura (motor separado del formato: el
número no se recalcula en el Excel). El layout es provisional; sólo se afirma sobre el CONTENIDO
estable (número, total del contrato), no sobre la estética que cambiará con la plantilla real.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from services.calculations.aiu import calcular_totales_cotizacion
from services.export.cotizacion import EmpresaCotizacion, render_cotizacion_excel


def _cotizacion():
    return SimpleNamespace(
        numero="PIM-007-2026", cliente_id=7, cliente_nombre="Alcaldía de La Estrella",
        nombre_obra="Pavimentación vía La Estrella", ubicacion="km 4", vigencia_dias=15,
        administracion_pct=Decimal("0.05"), imprevistos_pct=Decimal("0.03"),
        utilidad_pct=Decimal("0.04"), iva_sobre_utilidad_pct=Decimal("0.19"),
        condiciones="Validez 15 días\nNo incluye obras adicionales",
    )


def _items():
    return [
        SimpleNamespace(orden=1, descripcion="Base granular", unidad="m3",
                        cantidad=Decimal("1000"), valor_unitario=Decimal("10000")),
    ]


def _celdas(xlsx: bytes):
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(xlsx)).active
    return [c.value for row in ws.iter_rows() for c in row if c.value is not None]


def test_render_produce_xlsx_valido_con_total_del_contrato():
    cot = _cotizacion()
    items = _items()
    tot = calcular_totales_cotizacion(
        items, administracion_pct=cot.administracion_pct, imprevistos_pct=cot.imprevistos_pct,
        utilidad_pct=cot.utilidad_pct, iva_sobre_utilidad_pct=cot.iva_sobre_utilidad_pct,
    )
    xlsx = render_cotizacion_excel(cot, items, tot, EmpresaCotizacion())

    assert xlsx[:2] == b"PK"                       # zip magic de un .xlsx
    valores = _celdas(xlsx)
    assert "COTIZACIÓN No. PIM-007-2026" in valores
    assert "TOTAL CONTRATO" in valores
    assert "Alcaldía de La Estrella" in valores    # cliente enriquecido
    # el total del contrato (11.276.000) llega a la hoja como número (motor = fuente de verdad)
    assert float(tot.total) in valores


def test_render_sin_empresa_usa_defaults_y_no_falla_sin_condiciones():
    cot = _cotizacion()
    cot.condiciones = None
    cot.cliente_nombre = None                       # sin enriquecer → cae a "Cliente #<id>"
    xlsx = render_cotizacion_excel(cot, _items(), calcular_totales_cotizacion(
        _items(), administracion_pct=Decimal("0"), imprevistos_pct=Decimal("0"),
        utilidad_pct=Decimal("0"), iva_sobre_utilidad_pct=Decimal("0"),
    ))
    valores = _celdas(xlsx)
    assert any("Cliente #7" in str(v) for v in valores)
