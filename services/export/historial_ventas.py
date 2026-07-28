"""Export del libro de ventas a Excel — MOTOR separado del FORMATO.

`render_historial_excel` es una función PURA: recibe las filas ya cargadas y el total del período, y
devuelve los bytes del `.xlsx`. No toca la BD ni la red, así que se prueba generando bytes y
reabriéndolos con openpyxl (mismo trato que `services/export/cotizacion.py`).

⚠️ **`total_periodo` se recibe, NO se calcula sumando las filas.** Para fracciones y empaques el
POS guarda `precio_unitario = total_linea / cantidad` (`modules/ventas/service.py:492`), así que
`Σ(cantidad × precio_unitario)` puede diferir del total realmente cobrado en centavos. Si este motor
sumara los renglones, el Excel contradiría al KPI de la misma pantalla — y esa clase de diferencia
no se detecta mirando, se detecta cuando el contador reclama.

Las filas son duck-typed: basta con que expongan los atributos de
`modules.ventas.schemas.HistorialLinea`.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable, Protocol, Sequence

from core.config.timezone import to_co

_COP = '"$" #,##0'
_CANT = "#,##0.###"      # las cantidades son NUMERIC(12,3): 1.5 kg tiene que verse como 1,5

# El orden es el que pidió el dueño para la pantalla; el informe lo respeta para que la columna que
# busca esté donde la busca. FECHA va delante porque el Excel suele abarcar varios días.
_COLUMNAS: Sequence[tuple[str, int]] = (
    ("FECHA", 11), ("HORA", 8), ("VENTA #", 9), ("PRODUCTO", 38), ("CLIENTE", 24),
    ("CANT.", 9), ("VR UNITARIO", 14), ("MÉTODO", 20), ("VENDEDOR", 18), ("VR TOTAL", 14),
    ("ESTADO", 12),
)


class _FilaHistorial(Protocol):
    """Lo que el render lee de cada renglón (duck typing sobre `HistorialLinea`)."""

    consecutivo: int
    fecha: datetime
    estado: str
    producto: str
    cantidad: Decimal
    precio_unitario: Decimal
    total_linea: Decimal
    cliente: str
    vendedor: str
    metodo_pago: str
    pagos: list


def metodo_legible(fila: _FilaHistorial) -> str:
    """'mixto' no le dice nada a nadie: se expande a los métodos reales del cobro.

    Es la misma regla que aplica el tab en pantalla; vive acá para que el archivo que se manda por
    fuera no diga menos que lo que se ve adentro.
    """
    if fila.metodo_pago != "mixto" or not fila.pagos:
        return fila.metodo_pago
    return " + ".join(p.metodo for p in fila.pagos)


def render_historial_excel(
    filas: Iterable[_FilaHistorial],
    *,
    desde: date,
    hasta: date,
    total_periodo: Decimal,
    empresa: str = "",
) -> bytes:
    """Renderiza el libro de ventas del rango a `.xlsx` y devuelve sus bytes."""
    # Import perezoso: openpyxl solo se carga al exportar, no en el arranque de la app.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    filas = list(filas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    encabezado = f"Ventas del {desde:%d/%m/%Y} al {hasta:%d/%m/%Y}"
    if empresa:
        encabezado = f"{empresa} — {encabezado}"
    ws.append([encabezado])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    fila_titulos = 3
    ws.append([titulo for titulo, _ in _COLUMNAS])
    gris = PatternFill("solid", fgColor="EEEEEE")
    for i in range(1, len(_COLUMNAS) + 1):
        celda = ws.cell(row=fila_titulos, column=i)
        celda.font = Font(bold=True)
        celda.fill = gris
        celda.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = _COLUMNAS[i - 1][1]

    for fila in filas:
        # La hora se pasa a Colombia acá y no en el caller: el motor sigue puro y el informe nunca
        # sale en UTC, que es el error que nadie nota hasta que una venta de la noche cambia de día.
        local = to_co(fila.fecha)
        ws.append([
            local.strftime("%d/%m/%Y"),
            local.strftime("%H:%M"),
            fila.consecutivo,
            fila.producto,
            fila.cliente,
            float(fila.cantidad),
            float(fila.precio_unitario),
            metodo_legible(fila),
            fila.vendedor,
            float(fila.total_linea),
            fila.estado,
        ])

    for r in range(fila_titulos + 1, fila_titulos + 1 + len(filas)):
        ws.cell(row=r, column=6).number_format = _CANT
        ws.cell(row=r, column=7).number_format = _COP
        ws.cell(row=r, column=10).number_format = _COP

    # El pie lleva el total del PERÍODO recibido, no la suma de la columna: ver el aviso del módulo.
    pie = fila_titulos + len(filas) + 2
    ws.cell(row=pie, column=9, value="TOTAL DEL PERÍODO").font = Font(bold=True)
    celda_total = ws.cell(row=pie, column=10, value=float(total_periodo))
    celda_total.font = Font(bold=True)
    celda_total.number_format = _COP
    ws.cell(
        row=pie + 1, column=9,
        value="No incluye ventas anuladas",
    ).font = Font(italic=True, size=9)

    # Dos líneas que valen mucho al abrirlo: los títulos quedan fijos y cada columna filtrable.
    ws.freeze_panes = ws.cell(row=fila_titulos + 1, column=1)
    ws.auto_filter.ref = f"A{fila_titulos}:{get_column_letter(len(_COLUMNAS))}{fila_titulos + len(filas)}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
