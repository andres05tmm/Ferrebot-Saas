"""Export de la cotización AIU a Excel — MOTOR separado del FORMATO (plan PIM Fase 2, spec 03).

`render_cotizacion_excel` es una función PURA: recibe la cotización, sus ítems, el desglose AIU ya
calculado (`services.calculations.aiu.TotalesAIU` — NUNCA se recalcula aquí) y los datos de la empresa
emisora, y devuelve los bytes del `.xlsx`. No toca la BD ni la red: se puede testear generando bytes y
reabriéndolos con openpyxl.

⚠️ LAYOUT PROVISIONAL (BLOQUEO parcial). La plantilla real del cliente (refs PIM-010/011-2026) aún no
llegó; este layout es una aproximación fiel a lo descrito en la spec 03 (encabezado con empresa + NIT,
"COTIZACIÓN No. X", SEÑORES/OBRA, tabla ÍTEM|DESCRIPCIÓN|UND|CANT.|VR UNITARIO|VR TOTAL con borde y
cabecera gris, bloque de totales AIU alineado a la derecha con TOTAL CONTRATO resaltado, CONDICIONES en
viñetas, formato COP `$ #,##0`). Cuando llegue la plantilla, se ajusta SÓLO este archivo: el motor y la
firma no cambian. TODO(PDF): el export a PDF queda pospuesto a una fase posterior — requiere una
dependencia pesada (reportlab/weasyprint), fuera del alcance de esta fase (plan §5, "PDF profesional").

`cotizacion` e `items` son duck-typed: basta que expongan los atributos de columna del ORM
(`modules.obra.models.CotizacionObra`/`ItemCotizacionObra`). Si `cotizacion` trae `cliente_nombre`
(enriquecido por el caller) se usa en el "SEÑORES"; si no, cae a "Cliente #<id>" (provisional, hasta
resolver los datos fiscales completos del destinatario — también bloqueado por la plantilla real).
"""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Iterable, Protocol

from services.calculations.aiu import TotalesAIU

_COP = '"$" #,##0'  # formato de número COP (spec 03: `$ #,##0`)
_PCT_FMT = "0.##%"    # porcentaje AIU (fracción 0–1 → 5%)


class _LineaExcel(Protocol):
    """Atributos que el render lee de cada ítem (duck typing sobre el ORM)."""

    orden: int
    descripcion: str
    unidad: str
    cantidad: Decimal
    valor_unitario: Decimal


@dataclass(frozen=True, slots=True)
class EmpresaCotizacion:
    """Datos de la empresa EMISORA para el encabezado (provisionales para PIM hasta parametrizar)."""

    nombre: str = "Construcciones PIM S.A.S."
    nit: str | None = "901462287"
    direccion: str | None = None
    telefono: str | None = None


def render_cotizacion_excel(
    cotizacion,
    items: Iterable[_LineaExcel],
    totales: TotalesAIU,
    empresa: EmpresaCotizacion | None = None,
) -> bytes:
    """Renderiza la cotización AIU a `.xlsx` y devuelve sus bytes. Formato PROVISIONAL (ver módulo)."""
    # Import perezoso: openpyxl sólo se carga cuando se exporta (no en el arranque de la app).
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    empresa = empresa or EmpresaCotizacion()
    items = list(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"
    ws.sheet_view.showGridLines = False

    # Paleta / estilos (tokens locales; el look final llega con la plantilla real).
    gris_cabecera = PatternFill("solid", fgColor="4B5563")
    resalte_total = PatternFill("solid", fgColor="FEF3C7")
    borde = Border(*(Side(style="thin", color="D1D5DB"),) * 4)
    titulo_font = Font(bold=True, size=16, color="111827")
    sub_font = Font(bold=True, size=11, color="374151")
    th_font = Font(bold=True, color="FFFFFF")
    der = Alignment(horizontal="right")
    izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    centro = Alignment(horizontal="center", vertical="center")

    # Columnas: ÍTEM | DESCRIPCIÓN | UND | CANT. | VR UNITARIO | VR TOTAL
    anchos = (7, 46, 8, 12, 16, 18)
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Encabezado: empresa + NIT ───────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = empresa.nombre
    ws["A1"].font = titulo_font
    linea_nit = f"NIT {empresa.nit}" if empresa.nit else ""
    extra = " · ".join(filter(None, (linea_nit, empresa.direccion, empresa.telefono)))
    if extra:
        ws.merge_cells("A2:F2")
        ws["A2"] = extra
        ws["A2"].font = Font(size=9, color="6B7280")

    ws.merge_cells("A4:F4")
    ws["A4"] = f"COTIZACIÓN No. {getattr(cotizacion, 'numero', '')}"
    ws["A4"].font = Font(bold=True, size=13, color="111827")

    # SEÑORES (cliente) / OBRA (proyecto) / ubicación / vigencia
    cliente = getattr(cotizacion, "cliente_nombre", None) or f"Cliente #{getattr(cotizacion, 'cliente_id', '')}"
    fila = 5
    for etiqueta, valor in (
        ("SEÑORES:", cliente),
        ("OBRA:", getattr(cotizacion, "nombre_obra", "")),
        ("UBICACIÓN:", getattr(cotizacion, "ubicacion", None) or "—"),
        ("VIGENCIA:", f"{getattr(cotizacion, 'vigencia_dias', '')} días"),
    ):
        ws[f"A{fila}"] = etiqueta
        ws[f"A{fila}"].font = sub_font
        ws.merge_cells(f"B{fila}:F{fila}")
        ws[f"B{fila}"] = str(valor)
        fila += 1

    # ── Tabla de ítems ──────────────────────────────────────────────────────────
    fila += 1
    encabezados = ("ÍTEM", "DESCRIPCIÓN", "UND", "CANT.", "VR UNITARIO", "VR TOTAL")
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila, column=col, value=texto)
        celda.fill = gris_cabecera
        celda.font = th_font
        celda.alignment = centro
        celda.border = borde
    fila += 1

    for pos, item in enumerate(items, start=1):
        sub = Decimal(item.cantidad) * Decimal(item.valor_unitario)
        valores = (
            pos,
            item.descripcion,
            item.unidad,
            float(item.cantidad),
            float(item.valor_unitario),
            float(sub),
        )
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.border = borde
            if col == 1:
                celda.alignment = centro
            elif col == 2:
                celda.alignment = izq
            elif col == 3:
                celda.alignment = centro
            elif col == 4:
                celda.alignment = der
            else:  # VR UNITARIO / VR TOTAL
                celda.alignment = der
                celda.number_format = _COP
        fila += 1

    # ── Bloque de totales AIU (alineado a la derecha) ───────────────────────────
    fila += 1
    filas_totales = (
        ("Subtotal", totales.subtotal, getattr(cotizacion, "administracion_pct", None)),
        ("Administración", totales.administracion, getattr(cotizacion, "administracion_pct", None)),
        ("Imprevistos", totales.imprevistos, getattr(cotizacion, "imprevistos_pct", None)),
        ("Utilidad", totales.utilidad, getattr(cotizacion, "utilidad_pct", None)),
        ("IVA sobre utilidad", totales.iva_utilidad, getattr(cotizacion, "iva_sobre_utilidad_pct", None)),
    )
    # Etiqueta con el % entre paréntesis para A/I/U/IVA (el subtotal no lleva %).
    for i, (etiqueta, valor, pct) in enumerate(filas_totales):
        mostrar_pct = pct is not None and etiqueta != "Subtotal"
        et = f"{etiqueta} ({float(pct):.0%})" if mostrar_pct else etiqueta
        ws.cell(row=fila, column=5, value=et).alignment = der
        ws.cell(row=fila, column=5).font = sub_font if i == 0 else Font(size=10)
        c = ws.cell(row=fila, column=6, value=float(valor))
        c.number_format = _COP
        c.alignment = der
        fila += 1

    # TOTAL CONTRATO — resaltado
    ws.cell(row=fila, column=5, value="TOTAL CONTRATO").font = Font(bold=True, size=12, color="111827")
    ws.cell(row=fila, column=5).alignment = der
    total_cell = ws.cell(row=fila, column=6, value=float(totales.total))
    total_cell.number_format = _COP
    total_cell.font = Font(bold=True, size=12)
    total_cell.alignment = der
    for col in (5, 6):
        ws.cell(row=fila, column=col).fill = resalte_total
    fila += 2

    # ── Condiciones y observaciones (viñetas) ───────────────────────────────────
    condiciones = getattr(cotizacion, "condiciones", None)
    if condiciones:
        ws.cell(row=fila, column=1, value="CONDICIONES Y OBSERVACIONES").font = sub_font
        fila += 1
        for linea in str(condiciones).splitlines():
            linea = linea.strip()
            if not linea:
                continue
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
            texto = linea if linea.startswith(("•", "-", "*")) else f"• {linea}"
            celda = ws.cell(row=fila, column=1, value=texto)
            celda.alignment = izq
            celda.font = Font(size=9, color="374151")
            fila += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
