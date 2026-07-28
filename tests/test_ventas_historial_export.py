"""Export del libro de ventas a Excel — motor puro, sin BD ni red.

Se genera el archivo y se reabre con openpyxl (mismo trato que `test_cotizacion_obra_export.py`).
Se asserta sobre el CONTENIDO que le importa a quien abre el archivo, no sobre la estética.
"""
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from services.export.historial_ventas import metodo_legible, render_historial_excel

_UTC5 = timezone(timedelta(hours=-5))


def _fila(**kw):
    base = dict(
        consecutivo=1, fecha=datetime(2026, 7, 27, 14, 30, tzinfo=_UTC5), estado="completada",
        producto="Cemento gris 50kg", cantidad=Decimal("2"), precio_unitario=Decimal("20000"),
        total_linea=Decimal("40000"), cliente="Consumidor Final", vendedor="Andrés",
        metodo_pago="efectivo", pagos=[],
    )
    return SimpleNamespace(**{**base, **kw})


def _abrir(xlsx: bytes):
    return load_workbook(BytesIO(xlsx)).active


def test_una_fila_por_renglon_con_sus_columnas():
    ws = _abrir(render_historial_excel(
        [_fila(), _fila(consecutivo=2, producto="Varilla 1/2", total_linea=Decimal("5000"))],
        desde=date(2026, 7, 27), hasta=date(2026, 7, 27), total_periodo=Decimal("45000"),
    ))
    titulos = [c.value for c in ws[3]]
    assert titulos[:5] == ["FECHA", "HORA", "VENTA #", "PRODUCTO", "CLIENTE"]
    assert [ws.cell(row=r, column=4).value for r in (4, 5)] == ["Cemento gris 50kg", "Varilla 1/2"]


def test_el_total_del_pie_es_el_recibido_y_no_la_suma_de_las_lineas():
    """El invariante nº 1 del tab, en el archivo.

    Se le pasan filas cuyos renglones suman $40.000 y un `total_periodo` de $39.999 (el caso real de
    una fracción, donde el POS guarda `precio_unitario = total / cantidad`). Gana el parámetro: el
    Excel tiene que decir lo que de verdad se cobró, no lo que dan las multiplicaciones.
    """
    xlsx = render_historial_excel(
        [_fila()], desde=date(2026, 7, 27), hasta=date(2026, 7, 27),
        total_periodo=Decimal("39999"),
    )
    ws = _abrir(xlsx)
    valores = [ws.cell(row=r, column=10).value for r in range(1, ws.max_row + 1)]
    assert 39999 in valores
    assert sum(1 for v in valores if v == 40000) == 1      # solo el renglón, no el pie


def test_la_hora_sale_en_colombia_y_no_en_utc():
    """El error que nadie nota hasta que una venta de la noche aparece con la fecha del día siguiente.
    04:30 UTC del 28 son las 23:30 del 27 en Colombia."""
    ws = _abrir(render_historial_excel(
        [_fila(fecha=datetime(2026, 7, 28, 4, 30, tzinfo=timezone.utc))],
        desde=date(2026, 7, 27), hasta=date(2026, 7, 27), total_periodo=Decimal("40000"),
    ))
    assert ws.cell(row=4, column=1).value == "27/07/2026"
    assert ws.cell(row=4, column=2).value == "23:30"


def test_una_venta_mixta_dice_como_pagaron_de_verdad():
    """'mixto' no es un método de pago. El archivo que sale por fuera no puede decir menos que la
    pantalla."""
    fila = _fila(metodo_pago="mixto", pagos=[
        SimpleNamespace(metodo="efectivo", monto=Decimal("30000")),
        SimpleNamespace(metodo="transferencia", monto=Decimal("20000")),
    ])
    assert metodo_legible(fila) == "efectivo + transferencia"
    ws = _abrir(render_historial_excel(
        [fila], desde=date(2026, 7, 27), hasta=date(2026, 7, 27), total_periodo=Decimal("50000"),
    ))
    assert ws.cell(row=4, column=8).value == "efectivo + transferencia"


def test_una_venta_anulada_va_marcada_en_su_columna():
    """Sigue en el libro —es historia— pero el que abre el archivo tiene que poder descontarla."""
    ws = _abrir(render_historial_excel(
        [_fila(estado="anulada")], desde=date(2026, 7, 27), hasta=date(2026, 7, 27),
        total_periodo=Decimal("0"),
    ))
    assert ws.cell(row=4, column=11).value == "anulada"


def test_las_cantidades_fraccionadas_no_se_redondean_a_entero():
    """1,5 kg de varilla tiene que verse como 1,5. `NUMERIC(12,3)` mostrado como entero es una
    factura mal leída."""
    ws = _abrir(render_historial_excel(
        [_fila(cantidad=Decimal("1.5"))], desde=date(2026, 7, 27), hasta=date(2026, 7, 27),
        total_periodo=Decimal("40000"),
    ))
    assert ws.cell(row=4, column=6).value == 1.5
    assert "0.###" in ws.cell(row=4, column=6).number_format


def test_un_rango_sin_ventas_produce_un_archivo_valido():
    """Un mes sin ventas no puede reventar la descarga: sale el archivo con su encabezado y un cero."""
    ws = _abrir(render_historial_excel(
        [], desde=date(2026, 6, 1), hasta=date(2026, 6, 30), total_periodo=Decimal("0"),
    ))
    assert "01/06/2026" in ws["A1"].value and "30/06/2026" in ws["A1"].value
    assert [c.value for c in ws[3]][0] == "FECHA"
